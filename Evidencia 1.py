import numpy as np
from datetime import datetime, timedelta
from tabulate import tabulate
import csv
from openpyxl import Workbook
import re
import os

# Definición de tipos de datos
unidades_dtype = [('Clave', int), ('Rodada', int)]
clientes_dtype = [('Clave', int), ('Apellidos', 'U40'), ('Nombres', 'U40'), ('Telefono', 'U10')]
prestamos_dtype = [
    ('Folio', int), ('Clave Unidad', int), ('Clave Cliente', int), 
    ('Fecha Prestamo', 'U10'), ('Dias Prestamo', int), 
    ('Fecha Retorno', 'U10'), ('Fecha Retorno Efectiva', 'U10')
]

unidades = np.array([], dtype=unidades_dtype)
clientes = np.array([], dtype=clientes_dtype)
prestamos = np.array([], dtype=prestamos_dtype)

# Cargar datos
def cargar_datos():
    global unidades, clientes, prestamos
    try:
        if os.path.exists('unidades.csv'):
            unidades = np.genfromtxt('unidades.csv', delimiter=',', dtype=unidades_dtype, skip_header=1)
        if os.path.exists('clientes.csv'):
            clientes = np.genfromtxt('clientes.csv', delimiter=',', dtype=clientes_dtype, skip_header=1)
        if os.path.exists('prestamos.csv'):
            prestamos = np.genfromtxt('prestamos.csv', delimiter=',', dtype=prestamos_dtype, skip_header=1)
        
        # Validar que los campos sean los esperados
        if 'Fecha Retorno Efectiva' not in prestamos.dtype.names:
            raise ValueError("El campo 'Fecha Retorno Efectiva' no está presente en los datos cargados.")
        
    except IOError:
        print("No se encontraron archivos previos. Se iniciará con datos vacíos.")
    except ValueError as e:
        print(f"Error en la estructura de los datos cargados: {e}")
        # Reiniciar los datos vacíos si hay un error en los campos
        prestamos = np.array([], dtype=prestamos_dtype)

# Guardar datos
def guardar_datos():
    np.savetxt('unidades.csv', unidades, delimiter=',', header='Clave,Rodada', fmt='%d', comments='')
    np.savetxt('clientes.csv', clientes, delimiter=',', header='Clave,Apellidos,Nombres,Telefono', fmt='%s', comments='')
    np.savetxt('prestamos.csv', prestamos, delimiter=',', header='Folio,Clave Unidad,Clave Cliente,Fecha Prestamo,Dias Prestamo,Fecha Retorno,Fecha Retorno Efectiva', fmt='%s', comments='')

# Función para generar claves únicas
def generar_clave(arreglo):
    if arreglo.size > 0 and 'Clave' in arreglo.dtype.names:
        return arreglo['Clave'].max() + 1
    return 1

# Función para registrar unidades múltiples
def registrar_unidad():
    global unidades
    while True:
        clave = generar_clave(unidades)
        try:
            rodada = int(input("Ingrese la rodada de la bicicleta (20, 26, 29): "))
            if rodada not in [20, 26, 29]:
                raise ValueError("Rodada no válida. Debe ser 20, 26 o 29.")
            
            nueva_unidad = np.array([(clave, rodada)], dtype=unidades_dtype)
            unidades = np.append(unidades, nueva_unidad)
            print(f"Unidad registrada exitosamente con clave {clave}.")

            # Preguntar si desea registrar otra unidad
            otra = input("¿Desea registrar otra bicicleta? (s/n): ").strip().lower()
            if otra != 's':
                break
        except ValueError as e:
            print(e)


# Función para registrar un cliente
def registrar_cliente():
    global clientes
    while True:
        clave = generar_clave(clientes)

        # Validación para apellidos
        while True:
            apellidos = input("Ingrese los apellidos del cliente (máximo 40 caracteres): ").strip()[:40]
            if re.match("^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s]+$", apellidos):
                break
            else:
                print("Apellidos no válidos. Solo se permiten letras y espacios.")

        # Validación para nombres
        while True:
            nombres = input("Ingrese los nombres del cliente (máximo 40 caracteres): ").strip()[:40]
            if re.match("^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s]+$", nombres):
                break
            else:
                print("Nombres no válidos. Solo se permiten letras y espacios.")

        # Validación para teléfono
        while True:
            telefono = input("Ingrese el teléfono del cliente (10 dígitos, solo números): ").strip()
            if re.match("^\d{10}$", telefono):  # Acepta solo 10 dígitos
                break
            else:
                print("Teléfono no válido. Debe contener exactamente 10 dígitos numéricos.")

        nuevo_cliente = np.array([(clave, apellidos, nombres, telefono)], dtype=clientes_dtype)
        clientes = np.append(clientes, nuevo_cliente)
        print("Cliente registrado exitosamente.")
        break

# Función para registrar un préstamo
def registrar_prestamo():
    global prestamos
    try:
        # Obtener las claves de unidades que están actualmente prestadas y aún no han sido retornadas
        unidades_ocupadas = prestamos[prestamos['Fecha Retorno Efectiva'] == '']['Clave Unidad']

        # Filtrar unidades disponibles (aseguramos que sea un arreglo estructurado de NumPy)
        unidades_disponibles = unidades[np.isin(unidades['Clave'], unidades_ocupadas, invert=True)]

        # Verificar si hay unidades disponibles
        if unidades_disponibles.size == 0:
            print("No hay unidades disponibles para préstamo.")
            return  # Termina la función si no hay unidades

        # Validación de clave de cliente
        while True:
            try:
                clave_cliente = int(input("Ingrese la clave del cliente: "))
                if clave_cliente not in clientes['Clave']:
                    raise ValueError("Cliente no encontrado.")
                break
            except ValueError as e:
                print(f"Error: {e}. Ingrese una clave válida.")

        # Permitir registrar múltiples préstamos para el mismo cliente
        while True:
            print("\nUnidades Disponibles:")
            mostrar_reporte(unidades_disponibles, ['Clave', 'Rodada'])

            # Validación de clave de unidad
            while True:
                try:
                    clave_unidad = int(input("Ingrese la clave de la unidad a prestar (o 0 para finalizar): "))
                    if clave_unidad == 0:
                        print("Finalizando registro de préstamos.")
                        return  # Salir del bucle y terminar la función
                    if clave_unidad not in unidades_disponibles['Clave']:
                        raise ValueError("Unidad no disponible o no encontrada.")
                    break
                except ValueError as e:
                    print(f"Error: {e}. Ingrese una clave válida.")

            # Validación de fecha de préstamo
            while True:
                fecha_prestamo = input("Ingrese la fecha del préstamo (mm-dd-aaaa) o deje vacío para la fecha actual: ")
                if not fecha_prestamo:
                    fecha_prestamo = datetime.now().strftime("%m-%d-%Y")
                    break
                try:
                    fecha_prestamo_dt = datetime.strptime(fecha_prestamo, "%m-%d-%Y")
                    if fecha_prestamo_dt < datetime.now():
                        raise ValueError("La fecha de préstamo no puede ser anterior a la fecha actual.")
                    fecha_prestamo = fecha_prestamo_dt.strftime("%m-%d-%Y")
                    break
                except ValueError as e:
                    print(f"Error: {e}. Formato de fecha incorrecto.")

            # Validación de días de préstamo
            while True:
                try:
                    dias_prestamo = int(input("Ingrese la cantidad de días del préstamo (1-14): "))
                    if dias_prestamo < 1 or dias_prestamo > 14:
                        raise ValueError("Cantidad de días no válida. Debe ser entre 1 y 14.")
                    break
                except ValueError as e:
                    print(f"Error: {e}. Ingrese un número válido.")

            # Calcular la fecha de retorno
            fecha_retorno = (datetime.strptime(fecha_prestamo, "%m-%d-%Y") + timedelta(days=dias_prestamo)).strftime("%m-%d-%Y")
            folio = generar_clave(prestamos)

            # Registrar el nuevo préstamo
            nuevo_prestamo = np.array([(folio, clave_unidad, clave_cliente, fecha_prestamo, dias_prestamo, fecha_retorno, "")], dtype=prestamos_dtype)
            prestamos = np.append(prestamos, nuevo_prestamo)
            print("Préstamo registrado exitosamente.")

            # Actualizar unidades disponibles
            unidades_ocupadas = prestamos[prestamos['Fecha Retorno Efectiva'] == '']['Clave Unidad']
            unidades_disponibles = unidades[np.isin(unidades['Clave'], unidades_ocupadas, invert=True)]

    except Exception as e:
        print(f"Error al registrar el préstamo: {e}")

# Función para registrar la devolución de una bicicleta
def devolver_bicicleta():
    global prestamos
    try:
        # Filtrar préstamos activos (no devueltos)
        prestamos_activos = prestamos[prestamos['Fecha Retorno Efectiva'] == '']
        
        # Verificar si hay préstamos activos
        if prestamos_activos.size == 0:
            print("No hay unidades actualmente prestadas.")
            return
        
        # Mostrar préstamos activos
        print("\nPréstamos Activos:")
        mostrar_reporte(prestamos_activos, ['Folio', 'Clave Unidad', 'Clave Cliente', 'Fecha Prestamo', 'Dias Prestamo', 'Fecha Retorno', 'Fecha Retorno Efectiva'])
        
        # Validación de folio de préstamo
        while True:
            try:
                folio = int(input("Ingrese el folio del préstamo para registrar la devolución: "))
                # Validar que el folio existe en los préstamos activos
                if folio not in prestamos_activos['Folio']:
                    raise ValueError("Folio no encontrado o ya está registrado como devuelto.")
                break
            except ValueError as e:
                print(f"Error: {e}. Ingrese un folio válido.")
        
        # Actualizar la fecha de retorno efectiva
        while True:
            fecha_retorno_efectiva = input("Ingrese la fecha de retorno efectiva (mm-dd-aaaa) o deje vacío para la fecha actual: ")
            if not fecha_retorno_efectiva:
                fecha_retorno_efectiva = datetime.now().strftime("%m-%d-%Y")
                break
            try:
                fecha_retorno_efectiva_dt = datetime.strptime(fecha_retorno_efectiva, "%m-%d-%Y")
                # Obtener la fecha de retorno esperada
                fecha_retorno = datetime.strptime(prestamos[prestamos['Folio'] == folio]['Fecha Retorno'][0], "%m-%d-%Y")
                if fecha_retorno_efectiva_dt < fecha_retorno:
                    raise ValueError("La fecha de retorno efectiva no puede ser anterior a la fecha de retorno.")
                fecha_retorno_efectiva = fecha_retorno_efectiva_dt.strftime("%m-%d-%Y")
                break
            except ValueError as e:
                print(f"Error: {e}. Formato de fecha incorrecto.")
        
        # Actualizar el registro del préstamo con el folio correspondiente
        index = np.where(prestamos['Folio'] == folio)[0]
        if index.size > 0:
            prestamos[index[0]]['Fecha Retorno Efectiva'] = fecha_retorno_efectiva
            print("Devolución registrada exitosamente.")
        else:
            print("Error al actualizar la devolución.")

    except Exception as e:
        print(f"Error al registrar la devolución: {e}")

# Función para mostrar las devoluciones de préstamos
def mostrar_devoluciones():
    # Filtrar préstamos devueltos
    devoluciones = prestamos[prestamos['Fecha Retorno Efectiva'] != '']
    
    if devoluciones.size == 0:
        print("No hay devoluciones registradas.")
    else:
        print("\nReporte de Devoluciones:")
        mostrar_reporte(
            devoluciones,
            ['Folio', 'Clave Unidad', 'Clave Cliente', 'Fecha Prestamo', 'Dias Prestamo', 'Fecha Retorno', 'Fecha Retorno Efectiva']
        )


# Función para mostrar un reporte tabular de los registros
def mostrar_reporte(data, headers):
    print(tabulate(data, headers=headers, tablefmt='grid'))

def exportar_datos():
    try:
        print("\nOpciones de Exportación")
        print("1. Exportar a CSV")
        print("2. Exportar a Excel")
        print("0. Volver al menú principal")
        
        opcion = int(input("Seleccione una opción: "))
        
        if opcion == 0:
            return
        
        if opcion == 1:
            # Exportar a CSV
            guardar_datos()  # Asegúrate de que esta función se ejecute sin errores
            print("Datos exportados a CSV exitosamente.")
        
        elif opcion == 2:
            # Exportar a Excel
            print("Iniciando la exportación a Excel...")  # Mensaje de depuración
            wb = Workbook()
            
            # Crear hojas para cada tipo de datos
            ws_unidades = wb.active
            ws_unidades.title = "Unidades"
            ws_clientes = wb.create_sheet("Clientes")
            ws_prestamos = wb.create_sheet("Préstamos Activos")
            ws_devoluciones = wb.create_sheet("Devoluciones")

            # Encabezados de cada hoja
            ws_unidades.append(['Clave', 'Rodada'])
            ws_clientes.append(['Clave', 'Apellidos', 'Nombres', 'Teléfono'])
            ws_prestamos.append(['Folio', 'Clave Unidad', 'Clave Cliente', 'Fecha Préstamo', 'Días Préstamo', 'Fecha Retorno', 'Fecha Retorno Efectiva'])
            ws_devoluciones.append(['Folio', 'Clave Unidad', 'Clave Cliente', 'Fecha Préstamo', 'Días Préstamo', 'Fecha Retorno', 'Fecha Retorno Efectiva'])

            # Agregar datos a las hojas correspondientes
            print("Agregando datos de Unidades...")  # Mensaje de depuración
            for unidad in unidades:
                ws_unidades.append(unidad.tolist())
            
            print("Agregando datos de Clientes...")  # Mensaje de depuración
            for cliente in clientes:
                ws_clientes.append(cliente.tolist())
            
            # Dividir los préstamos en activos y devueltos
            print("Separando préstamos activos y devoluciones...")  # Mensaje de depuración
            prestamos_activos = prestamos[prestamos['Fecha Retorno Efectiva'] == '']
            devoluciones = prestamos[prestamos['Fecha Retorno Efectiva'] != '']
            
            # Agregar los préstamos activos y devueltos a sus respectivas hojas
            print("Agregando datos de Préstamos Activos...")  # Mensaje de depuración
            for prestamo in prestamos_activos:
                ws_prestamos.append(prestamo.tolist())
            
            print("Agregando datos de Devoluciones...")  # Mensaje de depuración
            for devolucion in devoluciones:
                ws_devoluciones.append(devolucion.tolist())
            
            # Guardar el archivo Excel
            excel_filename = "datos.xlsx"
            wb.save(excel_filename)
            print(f"Datos exportados a Excel exitosamente en '{excel_filename}'.")
        
        else:
            print("Opción no válida.")
    
    except Exception as e:
        print(f"Error al exportar los datos: {e}")

# Definición del menú principal y las opciones
def menu_principal():
    cargar_datos()
    opciones = {
        1: ('Registrar Unidad', registrar_unidad),
        2: ('Registrar Cliente', registrar_cliente),
        3: ('Registrar Préstamo', registrar_prestamo),
        4: ('Devolver Bicicleta', devolver_bicicleta),
        5: ('Mostrar Reporte de Unidades', lambda: mostrar_reporte(unidades, ['Clave', 'Rodada'])),
        6: ('Mostrar Reporte de Clientes', lambda: mostrar_reporte(clientes, ['Clave', 'Apellidos', 'Nombres', 'Telefono'])),
        7: ('Mostrar Reporte de Préstamos', lambda: mostrar_reporte(prestamos, ['Folio', 'Clave Unidad', 'Clave Cliente', 'Fecha Prestamo', 'Dias Prestamo', 'Fecha Retorno', 'Fecha Retorno Efectiva'])),
        8: ('Mostrar Reporte de Devoluciones', mostrar_devoluciones),
        9: ('Exportar Datos', exportar_datos),
        0: ('Salir', lambda: print("Saliendo del programa."))
    }
    while True:
        print("\nMenú Principal")
        for k, v in opciones.items():
            print(f"{k}. {v[0]}")
        try:
            opcion = int(input("Seleccione una opción: "))
            if opcion == 0:
                opciones[opcion][1]()
                break  # Salir del bucle y terminar el programa
            elif opcion in opciones:
                opciones[opcion][1]()
            else:
                print("Opción no válida. Intente de nuevo.")
        except ValueError:
            print("Entrada no válida. Ingrese un número.")

# Ejecutar el menú principal
if __name__ == "__main__":
    menu_principal()